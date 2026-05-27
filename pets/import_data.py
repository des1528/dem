"""Импорт ресурсов «Товары для животных»."""
import openpyxl, db, os, shutil
R = r"C:\demka\_extracted\pets\1-Ресурсы - Товары для животных"

def get_or_create(table, name):
    if not name: return None
    name = str(name).strip()
    r = db.fetch(f"SELECT ID FROM {table} WHERE Name=%s",(name,))
    return r[0]["ID"] if r else db.query(f"INSERT INTO {table}(Name) VALUES (%s)",(name,))

# --- Товары ---
wb = openpyxl.load_workbook(os.path.join(R, "Товар_import_Товары для животных.xlsx"))
ws = wb.active
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(header, row))
    art = d.get("Артикул")
    if not art: continue
    sid = get_or_create("Supplier", d.get("Поставщик"))
    mid = get_or_create("Manufacturer", d.get("Производитель"))
    cid = get_or_create("Category", d.get("Категория товара") or d.get("Категория"))
    db.query("""INSERT INTO Product(Article,Name,Unit,Price,SupplierID,ManufacturerID,CategoryID,Discount,Stock,Description,ImagePath)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (str(art), d.get("Наименование товара") or d.get("Наименование"),
              d.get("Единица измерения") or "шт.", float(d.get("Цена") or 0),
              sid, mid, cid, int(d.get("Действующая скидка") or 0),
              int(d.get("Кол-во на складе") or 0),
              d.get("Описание товара") or d.get("Описание"),
              str(d.get("Фото")) if d.get("Фото") else None))

# --- Пользователи ---
wb = openpyxl.load_workbook(os.path.join(R, "user_import.xlsx")); ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    role_name, fio, login, password, *_ = list(row) + [None]*4
    if not login: continue
    rid = get_or_create("Role", role_name)
    db.query("INSERT IGNORE INTO User(RoleID,FIO,Login,Password) VALUES (%s,%s,%s,%s)",
             (rid, fio, login, password))

# --- Пункты выдачи ---
wb = openpyxl.load_workbook(os.path.join(R, "Пункты выдачи_import.xlsx")); ws = wb.active
header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
addrs = [header[0]] + [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
for a in addrs: db.query("INSERT IGNORE INTO PickupPoint(Address) VALUES (%s)",(str(a),))

# --- Заказы ---
wb = openpyxl.load_workbook(os.path.join(R, "Заказ_import.xlsx")); ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    num, articles, odate, ddate, addr, fio, code, status, *_ = list(row) + [None]*8
    if not num: continue
    if isinstance(addr, int):
        pp = db.fetch("SELECT ID FROM PickupPoint LIMIT 1 OFFSET %s",(addr-1,))
        ppid = pp[0]["ID"] if pp else None
    else:
        r = db.fetch("SELECT ID FROM PickupPoint WHERE Address=%s",(addr,))
        ppid = r[0]["ID"] if r else db.query("INSERT INTO PickupPoint(Address) VALUES (%s)",(addr,))
    if not ppid: continue
    oid = db.query("""INSERT INTO `Order`(OrderDate,DeliveryDate,PickupPointID,ClientFIO,Code,Status)
                      VALUES (%s,%s,%s,%s,%s,%s)""",
                   (odate, ddate, ppid, fio, int(code or 0), status))
    if articles:
        parts = [p.strip() for p in str(articles).split(",")]
        for i in range(0, len(parts)-1, 2):
            p = db.fetch("SELECT ID FROM Product WHERE Article=%s",(parts[i],))
            if p:
                try: q = int(parts[i+1])
                except: q = 1
                db.query("INSERT INTO OrderItem(OrderID,ProductID,Qty) VALUES (%s,%s,%s)",(oid, p[0]["ID"], q))

print("Импорт завершён.")
