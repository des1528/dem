"""Импорт ресурсов магазина обуви.
Перед запуском: выполнить schema.sql и заполнить CFG в db.py.
"""
import openpyxl, db, os, shutil
from datetime import datetime

R = r"C:\demka\_extracted\m1\import"

def get_or_create(table, name):
    if not name: return None
    name = str(name).strip()
    r = db.fetch(f"SELECT ID FROM {table} WHERE Name=%s",(name,))
    return r[0]["ID"] if r else db.query(f"INSERT INTO {table}(Name) VALUES (%s)",(name,))

# --- Товары ---
wb = openpyxl.load_workbook(os.path.join(R,"Tovar.xlsx")); ws = wb.active
os.makedirs("images", exist_ok=True)
for art,name,unit,price,sup,manu,cat,disc,stock,desc,photo in list(ws.iter_rows(min_row=2, values_only=True)):
    if not art: continue
    sid = get_or_create("Supplier", sup)
    mid = get_or_create("Manufacturer", manu)
    cid = get_or_create("Category", cat)
    img = ""
    if photo:
        src = os.path.join(R, str(photo)); dst = os.path.join("images", str(photo))
        if os.path.exists(src) and not os.path.exists(dst): shutil.copy(src, dst)
        img = dst
    db.query("""INSERT INTO Product(Article,Name,Unit,Price,SupplierID,ManufacturerID,CategoryID,Discount,Stock,Description,ImagePath)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (art,name,unit,float(price or 0),sid,mid,cid,int(disc or 0),int(stock or 0),desc,img or None))

# --- Пользователи ---
wb = openpyxl.load_workbook(os.path.join(R,"user_import.xlsx")); ws = wb.active
for role_name, fio, login, password in list(ws.iter_rows(min_row=2, values_only=True)):
    if not login: continue
    r = db.fetch("SELECT ID FROM Role WHERE Name=%s",(role_name,))
    rid = r[0]["ID"] if r else db.query("INSERT INTO Role(Name) VALUES (%s)",(role_name,))
    db.query("INSERT IGNORE INTO User(RoleID,FIO,Login,Password) VALUES (%s,%s,%s,%s)",
             (rid, fio, login, password))

# --- Пункты выдачи ---
wb = openpyxl.load_workbook(os.path.join(R,"Пункты выдачи_import.xlsx")); ws = wb.active
# первая строка тоже адрес — она в заголовке
header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
addrs = [header[0]] + [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
for a in addrs:
    db.query("INSERT IGNORE INTO PickupPoint(Address) VALUES (%s)",(str(a),))

# --- Заказы ---
wb = openpyxl.load_workbook(os.path.join(R,"Заказ_import.xlsx")); ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    num, articles, odate, ddate, addr, fio, code, status, *_ = row
    if not num: continue
    if isinstance(addr, int):
        pp = db.fetch("SELECT ID FROM PickupPoint LIMIT 1 OFFSET %s",(addr-1,))
        ppid = pp[0]["ID"] if pp else None
    else:
        r = db.fetch("SELECT ID FROM PickupPoint WHERE Address=%s",(addr,))
        ppid = r[0]["ID"] if r else db.query("INSERT INTO PickupPoint(Address) VALUES (%s)",(addr,))
    if not ppid: continue
    oid = db.query("""INSERT INTO `Order`(OrderDate,DeliveryDate,PickupPointID,ClientFIO,Code,Status)
                      VALUES (%s,%s,%s,%s,%s,%s)""",
                   (odate, ddate, ppid, fio, int(code or 0), status))
    # «Артикул заказа» вида "А112Т4, 2, F635R4, 2" — пары (артикул, кол-во)
    if articles:
        parts = [p.strip() for p in str(articles).split(",")]
        for i in range(0, len(parts)-1, 2):
            art, qty = parts[i], parts[i+1]
            p = db.fetch("SELECT ID FROM Product WHERE Article=%s",(art,))
            if p: db.query("INSERT INTO OrderItem(OrderID,ProductID,Qty) VALUES (%s,%s,%s)",(oid, p[0]["ID"], int(qty)))

print("Импорт завершён.")
